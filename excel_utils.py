import streamlit as st
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font
from openpyxl.styles import PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo

def standardize(val):
    """
    将输入标准化为可比较的字符串：
    - 转字符串
    - 去除首尾空格
    - 去除单引号/双引号（英文和中文）
    - 统一半角/全角空格
    """
    if val is None:
        return ''
    
    val = str(val).strip()
    
    # 去掉包裹的引号（包括中英文单引号和双引号）
    val = val.strip('\'"“”‘’')  # 含中文引号
    
    # 替换全角空格为半角空格
    val = val.replace('\u3000', ' ')

    return val

def clean_df(df):
    df = df.fillna("")  # 将所有 NaN 替换为空字符串
    df = df.applymap(lambda x: str(x).strip() if isinstance(x, str) else x)  # 去除字符串中的前后空格
    return df



def adjust_column_width(writer, sheet_name, df):
    """
    自动调整 Excel 工作表中各列的宽度以适应内容长度。

    参数:
    - writer: pandas 的 ExcelWriter 对象
    - sheet_name: 要调整的工作表名称
    - df: 对应写入工作表的 DataFrame 数据
    """
    worksheet = writer.sheets[sheet_name]
    for idx, col in enumerate(df.columns, 1):
        # 获取该列中所有字符串长度的最大值
        max_content_len = df[col].astype(str).str.len().max()
        header_len = len(str(col))
        column_width = max(max_content_len, header_len) * 1.2 + 7
        worksheet.column_dimensions[get_column_letter(idx)].width = min(column_width, 50)

def merge_header_for_summary(ws, df, label_ranges):
    """
    给指定汇总列添加顶部合并行标题（如“安全库存”“未交订单”）

    参数:
    - ws: openpyxl worksheet
    - df: summary DataFrame
    - label_ranges: dict，键是标题文字，值是列名范围元组，如：
        {
            "安全库存": (" InvWaf", " InvPart"),
            "未交订单": ("总未交订单", "未交订单数量_2025-08")
        }
    """

    # 插入一行作为新表头（原表头往下挪）
    ws.insert_rows(1)
    header_row = list(df.columns)

    for label, (start_col_name, end_col_name) in label_ranges.items():
        if start_col_name not in header_row or end_col_name not in header_row:
            continue

        start_idx = header_row.index(start_col_name) + 1  # Excel index starts from 1
        end_idx = header_row.index(end_col_name) + 1

        col_letter_start = get_column_letter(start_idx)
        col_letter_end = get_column_letter(end_idx)

        merge_range = f"{col_letter_start}1:{col_letter_end}1"
        ws.merge_cells(merge_range)
        cell = ws[f"{col_letter_start}1"]
        cell.value = label
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(bold=True)

def mark_unmatched_keys_on_sheet(ws, unmatched_keys, wafer_col=1, spec_col=2, name_col=3):
    """
    在 openpyxl 工作表中标红未匹配的行（通过主键匹配），对空值/None做标准化处理。

    参数:
    - ws: openpyxl worksheet 对象
    - unmatched_keys: list of (晶圆品名, 规格, 品名) 元组
    - wafer_col, spec_col, name_col: 表示主键列在 sheet 中的列号（从1开始）
    """
    red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")

    unmatched_set = set(
        tuple(standardize(x) for x in key)
        for key in unmatched_keys
    )

    for row in range(2, ws.max_row + 1):  # 从第2行开始
        wafer = standardize(ws.cell(row=row, column=wafer_col).value)
        spec = standardize(ws.cell(row=row, column=spec_col).value)
        name = standardize(ws.cell(row=row, column=name_col).value)

        if (wafer, spec, name) in unmatched_set:
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).fill = red_fill


def mark_keys_on_sheet(ws, key_set, key_cols=(1, 2, 3)):
    """
    在工作表中标黄匹配 key_set 中的行，基于主键列匹配。

    参数:
    - ws: openpyxl worksheet
    - key_set: set of tuple，例如 {("晶圆品名", "规格", "品名"), ...}
    - key_cols: 表示主键所在的列号 (从1开始)，默认是 (1, 2, 3) 对应“晶圆品名”, “规格”, “品名”
    """
    from openpyxl.styles import PatternFill
    import re

    yellow_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")

    def standardize(val):
        if val is None:
            return ''
        val = str(val)
        val = val.replace('\u3000', ' ')  # 全角空格
        val = re.sub(r"[\"'‘’“”]", '', val)  # 引号
        return val.strip()

    # 标准化所有 key_set 中的值
    standardized_keys = set(tuple(standardize(x) for x in key) for key in key_set)

    # st.write(f"🟡 标黄匹配日志 - Sheet: {ws.title}")

    for row in range(2, ws.max_row + 1):  # 从第2行开始（跳过表头）
        key_raw = [ws.cell(row=row, column=col).value for col in key_cols]
        key = tuple(standardize(v) for v in key_raw)
        display_key = tuple(key_raw)  # 用原始值用于日志输出
        # st.write(f"第 {row} 行匹配尝试: {display_key}")
        if key in standardized_keys:
            # st.write(f"✅ 第 {row} 行匹配成功: {display_key}")
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).fill = yellow_fill
        # else:
            # st.write(f"❌ 第 {row} 行未匹配: {display_key}")

def merge_duplicate_product_names(summary_df: pd.DataFrame) -> pd.DataFrame:
    """
    合并 '汇总' 表中重复的品名（按品名分组），选用第一行的 晶圆品名 和 规格，合并其数值列。
    """
    # 确保必要列存在
    required_cols = ["晶圆品名", "规格", "品名"]
    for col in required_cols:
        if col not in summary_df.columns:
            raise ValueError(f"缺少必要列：{col}")

    # 识别数值列（排除主键列）
    value_cols = [col for col in summary_df.columns if col not in required_cols]

    # 分组合并数值列
    grouped = summary_df.groupby("品名", sort=False)

    merged_rows = []

    for name, group in grouped:
        if len(group) == 1:
            merged_rows.append(group.iloc[0])
        else:
            # 取第一行的 晶圆品名 和 规格
            base_row = group.iloc[0][required_cols].copy()
            summed_values = group[value_cols].apply(pd.to_numeric, errors="coerce").fillna(0).sum()
            merged_row = pd.concat([base_row, summed_values])
            merged_rows.append(merged_row)

    # 合并所有结果
    merged_df = pd.DataFrame(merged_rows)

    # 保证列顺序与原始一致
    return merged_df[summary_df.columns]

