import os
import re
import pandas as pd
import streamlit as st
from datetime import datetime, timedelta
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font
from openpyxl import load_workbook
from config import CONFIG, REVERSE_MAPPING
from excel_utils import (
    adjust_column_width, 
    merge_header_for_summary, 
    mark_unmatched_keys_on_sheet
)
from mapping_utils import apply_mapping_and_merge
from month_selector import process_history_columns
from summary import (
    merge_safety_inventory,
    append_unfulfilled_summary_columns,
    append_forecast_to_summary,
    merge_finished_inventory,
    append_product_in_progress
)

FIELD_MAPPINGS = {
    "unfulfilled_orders": {"规格": "规格", "品名": "品名", "晶圆品名": "晶圆品名"},
    "finished_products": {"规格": "产品规格", "品名": "产品品名", "晶圆品名": "晶圆型号"},
    "finished_inventory": {"规格": "规格", "品名": "品名", "晶圆品名": "WAFER品名"},
    "safety": {"规格": "OrderInformation", "品名": "ProductionNO.", "晶圆品名": "WaferID"},
    "forecast": {"规格": "产品型号", "品名": "ProductionNO.", "晶圆品名": "晶圆品名"}
}


class PivotProcessor:
    def process(self, uploaded_files: dict, output_buffer, additional_sheets: dict = None):
        with pd.ExcelWriter(output_buffer, engine="openpyxl") as writer:
            for filename, file_obj in uploaded_files.items():
                try:
                    df = pd.read_excel(file_obj)
                    st.write(filename)
                    config = CONFIG["pivot_config"].get(filename)
                    if not config:
                        st.warning(f"⚠️ 跳过未配置的文件：{filename}")
                        continue


                    sheet_name = filename[:30].replace(".xlsx", "")
                    st.write(f"📄 正在处理文件: `{filename}` → Sheet: `{sheet_name}`")


                    st.write(f"原始数据维度: {df.shape}")
                    st.dataframe(df.head(3))


                    # 日期处理
                    if "date_format" in config:
                        date_col = config["columns"]
                        df = self._process_date_column(df, date_col, config["date_format"])


                    # 映射替换（如果有）
                    st.write("additional_sheets")
                    st.write(additional_sheets)
                    if sheet_name in FIELD_MAPPINGS and "mapping" in (additional_sheets or {}):
                        mapping_df = additional_sheets["mapping"]


                        try:
                            mapping_df.columns = [
                                "旧规格", "旧品名", "旧晶圆品名",
                                "新规格", "新品名", "新晶圆品名",
                                "封装厂", "PC", "半成品"
                            ] + list(mapping_df.columns[9:])
                            st.success(f"✅ `{sheet_name}` 正在进行新旧料号替换...")
                        except Exception as e:
                            st.error(f"❌ `{sheet_name}` 替换前列名失败：{e}")
                            st.write("列名：", mapping_df.columns.tolist())
                            continue

                        
                        df = apply_mapping_and_merge(df, mapping_df, FIELD_MAPPINGS[sheet_name])


                    # 构建透视表
                    pivoted = self._create_pivot(df, config)
                    pivoted_display = pivoted.reset_index(drop=True)
                    if sheet_name == "finished_inventory":
                        df_finished = pivoted
                    elif sheet_name == "finished_products":
                        product_in_progress = pivoted

                    st.write(f"✅ Pivot 表创建成功，维度：{pivoted_display.shape}")
                    st.dataframe(pivoted_display.head(3))


                    excel_sheet_name = REVERSE_MAPPING.get(sheet_name, sheet_name)
                    pivoted.to_excel(writer, sheet_name=excel_sheet_name, index=False)
                    adjust_column_width(writer, excel_sheet_name, pivoted)

                    # 初始化未匹配变量，防止引用前未赋值
                    unmatched_safety = []
                    unmatched_unfulfilled = []
                    unmatched_forecast = []
                    unmatched_finished = []
                    unmatched_in_progress = []


                    # ✅ 如果当前是“未交订单”sheet，则拷贝前三列到新 sheet
                    if sheet_name == "赛卓-未交订单":
                        try:
                            # 提取前三列作为汇总基础
                            summary_preview = df[["晶圆品名", "规格", "品名"]].drop_duplicates().reset_index(drop=True)

                            # 追加安全库存信息
                            df_safety = additional_sheets["safety"]
                            summary_preview, unmatched_safety = merge_safety_inventory(summary_preview, df_safety)
                            st.success("✅ 已合并安全库存数据")
                            st.write(f"安全库存标红：{unmatched_safety}")


                            # 追加未交订单信息
                            summary_preview, unmatched_unfulfilled = append_unfulfilled_summary_columns(summary_preview, pivoted)
                            st.success("✅ 已合并未交订单数据")
                            st.write(f"未交订单标红：{unmatched_unfulfilled}")

                            # 追加预测信息
                            df_forecast = additional_sheets["赛卓-预测"]
                            df_forecast.columns = df_forecast.iloc[0]   # 第二行设为 header
                            df_forecast = df_forecast[1:].reset_index(drop=True)  # 删除第一行并重建索引
                            summary_preview, unmatched_forecast = append_forecast_to_summary(summary_preview, df_forecast)
                            st.success("✅ 已合并预测数据")
                            st.write(f"预测标红：{unmatched_forecast}")

                            # 追加成品库存信息
                            df_finished = apply_mapping_and_merge(df_finished, mapping_df, FIELD_MAPPINGS[sheet_name])
                            st.write(df_finished)
                            summary_preview, unmatched_finished = merge_finished_inventory(summary_preview, df_finished)
                            st.success("✅ 已合并成品库存")
                            st.write(f"库存信息标红：{unmatched_finished}")

                            # 追加成品在制信息
                            product_in_progress = apply_mapping_and_merge(product_in_progress, mapping_df, FIELD_MAPPINGS[sheet_name])
                            st.write(product_in_progress)
                            summary_preview, unmatched_in_progress = append_product_in_progress(summary_preview, product_in_progress, mapping_df)
                            st.success("✅ 已合并成品在制")
                            st.write(f"在制信息标红：{unmatched_in_progress}")



                            # 写入“汇总” sheet
                            summary_preview.to_excel(writer, sheet_name="汇总", index=False)
                            adjust_column_width(writer, "汇总", summary_preview)
                            st.success("✅ 已写入汇总Sheet")



                            # 打开 worksheet 进行格式化
                            ws = writer.sheets["汇总"]
                            header_row = list(summary_preview.columns)


                            # ✅ 找出所有“未交订单”相关列（顺序保留）
                            unfulfilled_cols = [col for col in header_row if (
                                col == "总未交订单" or 
                                col == "历史未交订单数量" or 
                                "未交订单数量" in col
                              )]
                            st.write(unfulfilled_cols)

                            # ✅ 找出所有“预测”相关列（顺序保留）
                            forecast_cols = [col for col in header_row if (
                                "预测" in col
                            )]
                            st.write(forecast_cols)

                            # ✅ 找出所有“成品库存”相关列（顺序保留）
                            finished_cols = [col for col in header_row if (
                                col == "数量_HOLD仓" or 
                                col == "数量_成品仓" or 
                                col == "数量_半成品仓"
                            )]
                            st.write(finished_cols)


                            merge_header_for_summary(
                                ws,
                                summary_preview,
                                {
                                    "安全库存": (" InvWaf", " InvPart"),
                                    "未交订单": (unfulfilled_cols[0], unfulfilled_cols[-1]),
                                    "预测": (forecast_cols[0], forecast_cols[-1]),
                                    "成品库存": (finished_cols[0], finished_cols[-1]),
                                    "成品在制": ("成品在制", "半成品在制")
                                 }
                            )



                        except Exception as e:
                            st.error(f"❌ 写入汇总失败: {e}")

                except Exception as e:
                    st.error(f"❌ 文件 `{filename}` 处理失败: {e}")

            # 写入新旧料号
            df_mapping = additional_sheets.get("mapping")
            if df_mapping is not None:
                df_mapping.to_excel(writer, sheet_name="赛卓-新旧料号", index=False)
                adjust_column_width(writer, "赛卓-新旧料号", df_mapping)



            # 写入附加 sheet（如预测、安全库存）
            if additional_sheets:
                for sheet_key, df in additional_sheets.items():
                    if sheet_key == "mapping":
                        continue
                    try:
                        sheet_name = REVERSE_MAPPING.get(sheet_key, sheet_key)  # 英文 ➜ 中文
                        st.write(f"📎 正在写入附加表：{sheet_name}，数据维度：{df.shape}")
                        df.to_excel(writer, sheet_name=sheet_name, index=False)
                        adjust_column_width(writer, sheet_name, df)
                    except Exception as e:
                        st.error(f"❌ 写入附加 Sheet `{sheet_key}` 失败: {e}")

            # 标记未匹配项
            try:
                ws = writer.sheets["赛卓-安全库存"]
                mark_unmatched_keys_on_sheet(ws, unmatched_safety, wafer_col=1, spec_col=3, name_col=5)
               
                ws = writer.sheets["赛卓-未交订单"]
                mark_unmatched_keys_on_sheet(ws, unmatched_unfulfilled, wafer_col=1, spec_col=2, name_col=3)
                
                ws = writer.sheets["赛卓-预测"]
                mark_unmatched_keys_on_sheet(ws, unmatched_forecast, wafer_col=3, spec_col=1, name_col=2)
                ws.delete_rows(2)  # 删除第 1 行
                
                ws = writer.sheets["赛卓-成品库存"]
                mark_unmatched_keys_on_sheet(ws, unmatched_finished, wafer_col=1, spec_col=2, name_col=3)
               
                ws = writer.sheets["赛卓-成品在制"]
                mark_unmatched_keys_on_sheet(ws, unmatched_in_progress, wafer_col=3, spec_col=4, name_col=5)

                ws = writer.sheets["赛卓-新旧料号"]
                ws.delete_rows(2)  # 删除第 1 行
                
                
                st.success("✅ 已完成未匹配项标记")

                # ✅ 所有写入完成后再加筛选器，避免被 to_excel 覆盖
                for sheet_name, ws in writer.sheets.items():
                    st.write(sheet_name)
                    # 如果第1行是你需要的 header，就添加筛选器
                    col_letter = get_column_letter(ws.max_column)
                    if sheet_name == "汇总":
                        ws.auto_filter.ref = f"A2:{col_letter}2"
                    else:
                        ws.auto_filter.ref = f"A1:{col_letter}1"

            except Exception as e:
                st.error(f"❌ 标记未匹配项失败: {e}")
        output_buffer.seek(0)

    def _process_date_column(self, df, date_col, date_format):
        if pd.api.types.is_numeric_dtype(df[date_col]):
            df[date_col] = df[date_col].apply(self._excel_serial_to_date)
        else:
            df[date_col] = pd.to_datetime(df[date_col], errors="coerce")

        new_col = f"{date_col}_年月"
        df[new_col] = df[date_col].dt.strftime(date_format)
        df[new_col] = df[new_col].fillna("未知日期")
        return df

    def _excel_serial_to_date(self, serial):
        try:
            return datetime(1899, 12, 30) + timedelta(days=float(serial))
        except:
            return pd.NaT

    def _create_pivot(self, df, config):
        config = config.copy()
        if "date_format" in config:
            config["columns"] = f"{config['columns']}_年月"


        pivoted = pd.pivot_table(
            df,
            index=config["index"],
            columns=config["columns"],
            values=config["values"],
            aggfunc=config["aggfunc"],
            fill_value=0
        )


        # 合并多级列名（如 (订单数量, 2024-05) → 订单数量_2024-05）
        pivoted.columns = [f"{col[0]}_{col[1]}" if isinstance(col, tuple) else str(col) for col in pivoted.columns]


        # 检查并处理重复列名
        if pd.Series(pivoted.columns).duplicated().any():
            from pandas.io.parsers import ParserBase
            original_cols = pivoted.columns
            deduped_cols = ParserBase({'names': original_cols})._maybe_dedup_names(original_cols)
            pivoted.columns = deduped_cols


        # 重置 index 以避免 to_excel 出错
        pivoted = pivoted.reset_index()

        # ✅ 仅对未交订单表触发历史数据合并
        if CONFIG.get("selected_month") and config.get("values") and "未交订单数量" in config.get("values"):
            st.info(f"📅 合并历史数据至：{CONFIG['selected_month']}")
            pivoted = process_history_columns(pivoted, config, CONFIG["selected_month"])
        return pivoted
